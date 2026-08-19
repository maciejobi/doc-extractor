"""
Eksport wyodrębnionych danych faktury do pliku Excel (.xlsx).

Buduje skoroszyt z trzema zakładkami:
  1. Podsumowanie - KPI (kwoty, liczba pozycji), rozbicie VAT, wykres kołowy.
  2. Pozycje - tabela pozycji z formułami (wartość = ilość × cena) i sumą.
  3. Dane faktury - pola nagłówkowe (sprzedawca, nabywca, daty, konto...).

Kwoty w tabeli pozycji liczone FORMUŁAMI Excela (nie zahardkodowane), więc
arkusz przelicza się po edycji. Formatowanie: Arial, granatowe nagłówki.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter


NAVY = "1B2A4A"
BLUE = "2D5B9A"
LIGHT = "EEF2F9"
MID = "D8E0EE"
WHITE = "FFFFFF"

_thin = Side(style="thin", color="C9D3E4")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
FONT = "Arial"

# --- Tłumaczenia etykiet Excela (PL/EN/DE/FR) ---
I18N = {
    "pl": {
        "summary_title": "Podsumowanie faktury", "nr": "Nr", "kpi": "WSKAŹNIKI",
        "net": "Suma netto", "vat_rate": "Stawka VAT", "vat": "Suma VAT",
        "gross": "Kwota brutto", "count": "Liczba pozycji",
        "vat_breakdown": "ROZBICIE VAT WG STAWEK", "rate": "Stawka VAT",
        "col_net": "Netto", "col_vat": "VAT", "col_gross": "Brutto",
        "no_rate_data": "brak danych o stawkach w pozycjach",
        "chart_title": "Udział netto wg stawki VAT",
        "items_title": "Pozycje faktury", "lp": "Lp.", "desc": "Opis",
        "qty": "Ilość", "unit_price": "Cena jedn.", "value_net": "Wartość netto", "value_gross": "Wartość brutto", "total_row": "RAZEM",
        "data_title": "Dane faktury", "sheet_sum": "Podsumowanie",
        "sheet_items": "Pozycje", "sheet_data": "Dane faktury",
    },
    "en": {
        "summary_title": "Invoice summary", "nr": "No.", "kpi": "KEY FIGURES",
        "net": "Net amount", "vat_rate": "VAT rate", "vat": "VAT amount",
        "gross": "Gross amount", "count": "Line items",
        "vat_breakdown": "VAT BREAKDOWN BY RATE", "rate": "VAT rate",
        "col_net": "Net", "col_vat": "VAT", "col_gross": "Gross",
        "no_rate_data": "no rate data in line items",
        "chart_title": "Net share by VAT rate",
        "items_title": "Line items", "lp": "No.", "desc": "Description",
        "qty": "Qty", "unit_price": "Unit price", "value_net": "Net amount", "value_gross": "Gross amount", "total_row": "TOTAL",
        "data_title": "Invoice data", "sheet_sum": "Summary",
        "sheet_items": "Line items", "sheet_data": "Invoice data",
    },
    "de": {
        "summary_title": "Rechnungsübersicht", "nr": "Nr.", "kpi": "KENNZAHLEN",
        "net": "Nettobetrag", "vat_rate": "MwSt-Satz", "vat": "MwSt-Betrag",
        "gross": "Bruttobetrag", "count": "Positionen",
        "vat_breakdown": "MwSt-AUFSCHLÜSSELUNG NACH SATZ", "rate": "MwSt-Satz",
        "col_net": "Netto", "col_vat": "MwSt", "col_gross": "Brutto",
        "no_rate_data": "keine Satzdaten in Positionen",
        "chart_title": "Nettoanteil nach MwSt-Satz",
        "items_title": "Rechnungspositionen", "lp": "Nr.", "desc": "Beschreibung",
        "qty": "Menge", "unit_price": "Einzelpreis", "value_net": "Nettobetrag", "value_gross": "Bruttobetrag", "total_row": "GESAMT",
        "data_title": "Rechnungsdaten", "sheet_sum": "Übersicht",
        "sheet_items": "Positionen", "sheet_data": "Rechnungsdaten",
    },
    "fr": {
        "summary_title": "Résumé de la facture", "nr": "N°", "kpi": "INDICATEURS",
        "net": "Montant HT", "vat_rate": "Taux TVA", "vat": "Montant TVA",
        "gross": "Montant TTC", "count": "Lignes",
        "vat_breakdown": "VENTILATION TVA PAR TAUX", "rate": "Taux TVA",
        "col_net": "HT", "col_vat": "TVA", "col_gross": "TTC",
        "no_rate_data": "pas de données de taux dans les lignes",
        "chart_title": "Part HT par taux de TVA",
        "items_title": "Lignes de facture", "lp": "N°", "desc": "Description",
        "qty": "Qté", "unit_price": "Prix unit.", "value_net": "Montant HT", "value_gross": "Montant TTC", "total_row": "TOTAL",
        "data_title": "Données de la facture", "sheet_sum": "Résumé",
        "sheet_items": "Lignes", "sheet_data": "Données facture",
    },
}


LABELS = {
    "invoice_number": "Numer faktury", "issue_date": "Data wystawienia",
    "service_date": "Data wykonania", "due_date": "Termin płatności",
    "seller_name": "Sprzedawca", "seller": "Sprzedawca", "seller_nip": "NIP sprzedawcy",
    "seller_address": "Adres sprzedawcy", "buyer_name": "Nabywca", "buyer": "Nabywca",
    "buyer_nip": "NIP nabywcy", "buyer_address": "Adres nabywcy",
    "net_amount": "Suma netto", "tax_amount": "Suma VAT", "total_amount": "Kwota brutto",
    "currency": "Waluta", "payment_method": "Metoda płatności",
    "bank_account": "Numer konta", "bank_bic": "BIC",
    "reference_numbers": "Numery referencyjne", "email": "E-mail", "vat_id": "NIP UE",
}

# etykiety pól nagłówkowych w pozostałych językach
FIELD_LABELS = {
    "pl": LABELS,
    "en": {
        "invoice_number": "Invoice number", "issue_date": "Issue date",
        "service_date": "Service date", "due_date": "Due date",
        "seller_name": "Seller", "seller": "Seller", "seller_nip": "Seller tax ID",
        "seller_address": "Seller address", "buyer_name": "Buyer", "buyer": "Buyer",
        "buyer_nip": "Buyer tax ID", "buyer_address": "Buyer address",
        "currency": "Currency", "payment_method": "Payment method",
        "bank_account": "Bank account", "bank_bic": "BIC",
        "reference_numbers": "Reference numbers", "email": "E-mail", "vat_id": "EU VAT ID",
    },
    "de": {
        "invoice_number": "Rechnungsnummer", "issue_date": "Rechnungsdatum",
        "service_date": "Leistungsdatum", "due_date": "Fälligkeitsdatum",
        "seller_name": "Verkäufer", "seller": "Verkäufer", "seller_nip": "USt-IdNr. Verkäufer",
        "seller_address": "Adresse Verkäufer", "buyer_name": "Käufer", "buyer": "Käufer",
        "buyer_nip": "USt-IdNr. Käufer", "buyer_address": "Adresse Käufer",
        "currency": "Währung", "payment_method": "Zahlungsart",
        "bank_account": "Bankkonto", "bank_bic": "BIC",
        "reference_numbers": "Referenznummern", "email": "E-Mail", "vat_id": "USt-IdNr.",
    },
    "fr": {
        "invoice_number": "Numéro de facture", "issue_date": "Date d'émission",
        "service_date": "Date de prestation", "due_date": "Date d'échéance",
        "seller_name": "Vendeur", "seller": "Vendeur", "seller_nip": "N° TVA vendeur",
        "seller_address": "Adresse vendeur", "buyer_name": "Client", "buyer": "Client",
        "buyer_nip": "N° TVA client", "buyer_address": "Adresse client",
        "currency": "Devise", "payment_method": "Mode de paiement",
        "bank_account": "Compte bancaire", "bank_bic": "BIC",
        "reference_numbers": "Numéros de référence", "email": "E-mail", "vat_id": "N° TVA",
    },
}
_ALIASES = {"seller_name": "seller", "buyer_name": "buyer"}

HEADER_FIELDS = [
    "invoice_number", "issue_date", "service_date", "due_date",
    "seller_name", "seller_nip", "seller_address",
    "buyer_name", "buyer_nip", "buyer_address",
    "currency", "payment_method", "bank_account", "bank_bic",
    "reference_numbers", "email", "vat_id",
]


def _get(fields: dict, key: str):
    if fields.get(key) not in (None, "", []):
        return fields[key]
    alias = _ALIASES.get(key)
    if alias and fields.get(alias) not in (None, "", []):
        return fields[alias]
    return None


def _num(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (ValueError, TypeError):
        return None


def build_xlsx(data: dict[str, Any], lang: str = "pl") -> bytes:
    if lang not in I18N:
        lang = "pl"
    L = I18N[lang]
    flabels = FIELD_LABELS.get(lang, FIELD_LABELS["pl"])

    wb = Workbook()
    fields = data.get("fields", {}) or {}
    items = data.get("line_items", []) or []
    currency = _get(fields, "currency") or "PLN"

    ws_sum = wb.active
    ws_sum.title = L["sheet_sum"]
    ws_items = wb.create_sheet(L["sheet_items"])
    ws_data = wb.create_sheet(L["sheet_data"])

    _build_items(ws_items, items, currency, L)
    _build_summary(ws_sum, fields, items, currency, len(items), L)
    _build_data(ws_data, fields, L, flabels)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# --- Zakładka 1: Podsumowanie ---------------------------------------------

def _build_summary(ws, fields, items, currency, n_items, L):
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [3, 22, 22, 18, 18, 14]):
        ws.column_dimensions[col].width = w

    ws["B2"] = L["summary_title"]
    ws["B2"].font = Font(name=FONT, size=16, bold=True, color=NAVY)
    inv = _get(fields, "invoice_number")
    if inv:
        ws["B3"] = f'{L["nr"]} {inv}'
        ws["B3"].font = Font(name=FONT, size=10, italic=True, color="7A869A")

    money_fmt = f'#,##0.00 "{currency}"'

    # KPI
    row = 5
    ws[f"B{row}"] = L["kpi"]
    ws[f"B{row}"].font = Font(name=FONT, size=10, bold=True, color=BLUE)
    row += 1

    net = _num(_get(fields, "net_amount"))
    tax = _num(_get(fields, "tax_amount"))
    total = _num(_get(fields, "total_amount"))

    # stawka VAT do KPI: z pozycji (jeśli jest) lub policzona z tax/net
    vat_rate_display = _dominant_vat_rate(items)
    if vat_rate_display is None and net and tax:
        vat_rate_display = round(tax / net * 100)  # np. 23

    kpis = [
        (L["net"], net, money_fmt),
        (L["vat_rate"], f"{vat_rate_display}%" if vat_rate_display is not None else "—", "@"),
        (L["vat"], tax, money_fmt),
        (L["gross"], total, money_fmt),
        (L["count"], n_items, "0"),
    ]
    col = 2
    for label, val, fmt in kpis:
        c_lab = ws.cell(row=row, column=col, value=label)
        c_lab.font = Font(name=FONT, size=9, color="5A6782")
        c_lab.fill = PatternFill("solid", fgColor=LIGHT)
        c_lab.border = BORDER
        c_lab.alignment = Alignment(horizontal="center")
        c_val = ws.cell(row=row + 1, column=col, value=val if val is not None else "—")
        c_val.font = Font(name=FONT, size=13, bold=True, color=NAVY)
        c_val.border = BORDER
        c_val.alignment = Alignment(horizontal="center")
        if val is not None and isinstance(val, (int, float)):
            c_val.number_format = fmt
        col += 1

    # rozbicie VAT wg stawek
    row += 3
    ws[f"B{row}"] = L["vat_breakdown"]
    ws[f"B{row}"].font = Font(name=FONT, size=10, bold=True, color=BLUE)
    row += 1
    for i, h in enumerate([L["rate"], L["col_net"], L["col_vat"], L["col_gross"]]):
        c = ws.cell(row=row, column=2 + i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    header_row = row
    row += 1

    vat_groups = _group_by_vat(items)
    data_start = row
    if vat_groups:
        for rate, (net_sum, vat_sum) in sorted(vat_groups.items()):
            ws.cell(row=row, column=2, value=rate).border = BORDER
            ws.cell(row=row, column=2).font = Font(name=FONT, size=10)
            for c_idx, v in [(3, net_sum), (4, vat_sum), (5, net_sum + vat_sum)]:
                cc = ws.cell(row=row, column=c_idx, value=round(v, 2))
                cc.number_format = money_fmt
                cc.border = BORDER
                cc.font = Font(name=FONT, size=10)
            row += 1
    else:
        ws.cell(row=row, column=2, value=L["no_rate_data"]).font = \
            Font(name=FONT, size=9, italic=True, color="9AA5B8")
        row += 1
    data_end = row - 1

    # wykres kołowy: udział netto wg stawki VAT
    if vat_groups:
        pie = PieChart()
        pie.title = L["chart_title"]
        labels_ref = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
        data_ref = Reference(ws, min_col=3, min_row=header_row, max_row=data_end)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        pie.height = 7
        pie.width = 11
        ws.add_chart(pie, f"B{row + 1}")


def _dominant_vat_rate(items):
    """Zwraca najczęstszą stawkę VAT z pozycji jako liczbę (np. 23) lub None."""
    from collections import Counter
    rates = []
    for it in items:
        if not isinstance(it, dict):
            continue
        r = _parse_rate((it.get("vat_rate") or "").strip())
        if r is not None:
            rates.append(round(r * 100))
    if not rates:
        return None
    return Counter(rates).most_common(1)[0][0]


def _group_by_vat(items) -> dict:
    """
    Grupuje pozycje wg stawki VAT: {stawka: (suma_netto, suma_vat)}.
    Wartość pozycji traktujemy jako NETTO (tak jest na większości faktur)
    i doliczamy VAT wg stawki. Gdy brak stawki - VAT = 0.
    """
    groups: dict[str, list] = {}
    for it in items:
        if not isinstance(it, dict):
            continue
        rate = (it.get("vat_rate") or "").strip()
        value = _num(it.get("total"))
        if value is None:
            continue
        key = rate or "—"
        rate_num = _parse_rate(rate)
        if rate_num is not None:
            net = value              # wartość pozycji = netto
            vat = value * rate_num   # VAT doliczony wg stawki
        else:
            net, vat = value, 0.0
        groups.setdefault(key, [0.0, 0.0])
        groups[key][0] += net
        groups[key][1] += vat
    return {k: (v[0], v[1]) for k, v in groups.items()}


def _parse_rate(rate: str):
    if not rate:
        return None
    m = "".join(ch for ch in rate if ch.isdigit() or ch == ".")
    try:
        return float(m) / 100 if m else None
    except ValueError:
        return None


# --- Zakładka 2: Pozycje ---------------------------------------------------

def _build_items(ws, items, currency, L):
    ws.sheet_view.showGridLines = False
    ws["A1"] = L["items_title"]
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)

    headers = [L["lp"], L["desc"], L["qty"], L["unit_price"], L["rate"],
               L["value_net"], L["value_gross"]]
    widths = [5, 30, 8, 14, 10, 15, 15]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center" if i != 2 else "left")
        ws.column_dimensions[get_column_letter(i)].width = w

    money_fmt = f'#,##0.00 "{currency}"'
    first_row = 4
    r = first_row
    total_net = 0.0
    total_gross = 0.0
    for idx, it in enumerate(items, start=1):
        if not isinstance(it, dict):
            continue
        qty = _num(it.get("quantity"))
        price = _num(it.get("unit_price"))
        rate_num = _parse_rate((it.get("vat_rate") or "").strip())
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=it.get("description", ""))
        ws.cell(row=r, column=3, value=qty if qty is not None else "")
        ws.cell(row=r, column=4, value=price if price is not None else "")
        ws.cell(row=r, column=5, value=it.get("vat_rate", "") or "")
        # wartość netto: ilość × cena, a jeśli brak - podana wartość z odczytu
        net_val = None
        if qty is not None and price is not None:
            net_val = round(qty * price, 2)
        else:
            net_val = _num(it.get("total"))
        # wartość brutto: netto × (1 + stawka VAT); bez stawki = netto
        gross_val = None
        if net_val is not None:
            gross_val = round(net_val * (1 + rate_num), 2) if rate_num is not None else net_val
        ws.cell(row=r, column=6, value=net_val if net_val is not None else "")
        ws.cell(row=r, column=7, value=gross_val if gross_val is not None else "")
        if net_val is not None:
            total_net += net_val
        if gross_val is not None:
            total_gross += gross_val
        for col in range(1, 8):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            cell.font = Font(name=FONT, size=10)
            if col in (4, 6, 7):
                cell.number_format = money_fmt
            if col == 2:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center" if col == 1 else "right")
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
        r += 1
    last_row = r - 1

    if items:
        c_lab = ws.cell(row=r, column=5, value=L["total_row"])
        c_lab.font = Font(name=FONT, bold=True, color=NAVY)
        c_lab.alignment = Alignment(horizontal="right")
        for col, val in [(6, total_net), (7, total_gross)]:
            c_sum = ws.cell(row=r, column=col, value=round(val, 2))
            c_sum.font = Font(name=FONT, bold=True, color=NAVY)
            c_sum.number_format = money_fmt
            c_sum.fill = PatternFill("solid", fgColor=MID)
            c_sum.border = BORDER


# --- Zakładka 3: Dane faktury ---------------------------------------------

def _build_data(ws, fields, L, flabels):
    ws.sheet_view.showGridLines = False
    ws["A1"] = L["data_title"]
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 46

    r = 3
    for key in HEADER_FIELDS:
        val = _get(fields, key)
        if val in (None, "", []):
            continue
        lab = ws.cell(row=r, column=1, value=flabels.get(key, key))
        lab.font = Font(name=FONT, size=10, bold=True, color=NAVY)
        lab.fill = PatternFill("solid", fgColor=LIGHT)
        lab.border = BORDER
        v = ws.cell(row=r, column=2, value=_fmt(val))
        v.font = Font(name=FONT, size=10)
        v.border = BORDER
        v.alignment = Alignment(wrap_text=True, vertical="center")
        r += 1


def _fmt(v: Any) -> str:
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return str(v)
