"""
Eksport zbiorczy WIELU faktur do jednego pliku Excel i jednego CSV pod Power BI.

Excel (zbiorczy) ma trzy zakładki:
  1. Podsumowanie - globalne KPI (suma netto/VAT/brutto wszystkich faktur,
     liczba faktur, liczba pozycji) + wykres udziału kwot wg kontrahenta.
  2. Faktury - lista: jeden wiersz = jedna faktura (nr, data, kontrahent, kwoty).
  3. Pozycje - wszystkie pozycje ze wszystkich faktur (z numerem faktury).

CSV (zbiorczy) - płaski plik: jeden wiersz = jedna pozycja z KAŻDEJ faktury,
gotowy pod Power BI do analizy wielu faktur naraz.

Języki: PL/EN/DE/FR (jak reszta aplikacji).
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from excel_export import _get, _num, NAVY, BLUE, LIGHT, MID, WHITE, FONT, BORDER

# tłumaczenia nagłówków zbiorczych
B_I18N = {
    "pl": {
        "title": "Zestawienie faktur", "kpi": "WSKAŹNIKI ZBIORCZE",
        "count_inv": "Liczba faktur", "sum_net": "Suma netto",
        "sum_vat": "Suma VAT", "sum_gross": "Suma brutto", "count_items": "Liczba pozycji",
        "by_seller": "KWOTY WG SPRZEDAWCY", "seller": "Sprzedawca", "gross": "Brutto",
        "chart_title": "Udział brutto wg sprzedawcy",
        "invoices_title": "Lista faktur", "items_title": "Wszystkie pozycje",
        "nr": "Nr faktury", "date": "Data", "buyer": "Nabywca",
        "net": "Netto", "vat": "VAT", "currency": "Waluta", "file": "Plik",
        "desc": "Opis", "qty": "Ilość", "unit_price": "Cena jedn.",
        "vat_rate": "Stawka VAT", "value_net": "Wartość netto",
        "sheet_sum": "Podsumowanie", "sheet_inv": "Faktury", "sheet_items": "Pozycje",
        "total_row": "RAZEM",
    },
    "en": {
        "title": "Invoice overview", "kpi": "AGGREGATE FIGURES",
        "count_inv": "Invoices", "sum_net": "Total net",
        "sum_vat": "Total VAT", "sum_gross": "Total gross", "count_items": "Line items",
        "by_seller": "AMOUNTS BY SELLER", "seller": "Seller", "gross": "Gross",
        "chart_title": "Gross share by seller",
        "invoices_title": "Invoice list", "items_title": "All line items",
        "nr": "Invoice no.", "date": "Date", "buyer": "Buyer",
        "net": "Net", "vat": "VAT", "currency": "Currency", "file": "File",
        "desc": "Description", "qty": "Qty", "unit_price": "Unit price",
        "vat_rate": "VAT rate", "value_net": "Net amount",
        "sheet_sum": "Summary", "sheet_inv": "Invoices", "sheet_items": "Line items",
        "total_row": "TOTAL",
    },
    "de": {
        "title": "Rechnungsübersicht", "kpi": "GESAMTKENNZAHLEN",
        "count_inv": "Rechnungen", "sum_net": "Netto gesamt",
        "sum_vat": "MwSt gesamt", "sum_gross": "Brutto gesamt", "count_items": "Positionen",
        "by_seller": "BETRÄGE NACH VERKÄUFER", "seller": "Verkäufer", "gross": "Brutto",
        "chart_title": "Bruttoanteil nach Verkäufer",
        "invoices_title": "Rechnungsliste", "items_title": "Alle Positionen",
        "nr": "Rechnungsnr.", "date": "Datum", "buyer": "Käufer",
        "net": "Netto", "vat": "MwSt", "currency": "Währung", "file": "Datei",
        "desc": "Beschreibung", "qty": "Menge", "unit_price": "Einzelpreis",
        "vat_rate": "MwSt-Satz", "value_net": "Nettobetrag",
        "sheet_sum": "Übersicht", "sheet_inv": "Rechnungen", "sheet_items": "Positionen",
        "total_row": "GESAMT",
    },
    "fr": {
        "title": "Récapitulatif des factures", "kpi": "INDICATEURS GLOBAUX",
        "count_inv": "Factures", "sum_net": "Total HT",
        "sum_vat": "Total TVA", "sum_gross": "Total TTC", "count_items": "Lignes",
        "by_seller": "MONTANTS PAR VENDEUR", "seller": "Vendeur", "gross": "TTC",
        "chart_title": "Part TTC par vendeur",
        "invoices_title": "Liste des factures", "items_title": "Toutes les lignes",
        "nr": "N° facture", "date": "Date", "buyer": "Client",
        "net": "HT", "vat": "TVA", "currency": "Devise", "file": "Fichier",
        "desc": "Description", "qty": "Qté", "unit_price": "Prix unit.",
        "vat_rate": "Taux TVA", "value_net": "Montant HT",
        "sheet_sum": "Résumé", "sheet_inv": "Factures", "sheet_items": "Lignes",
        "total_row": "TOTAL",
    },
}


def _seller_of(fields):
    return _get(fields, "seller_name") or _get(fields, "seller") or "—"


def build_batch_xlsx(invoices: list[dict], lang: str = "pl") -> bytes:
    """invoices: lista wyników to_dict() z wielu faktur."""
    if lang not in B_I18N:
        lang = "pl"
    L = B_I18N[lang]
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = L["sheet_sum"]
    ws_inv = wb.create_sheet(L["sheet_inv"])
    ws_items = wb.create_sheet(L["sheet_items"])

    _batch_invoices(ws_inv, invoices, L)
    _batch_items(ws_items, invoices, L)
    _batch_summary(ws_sum, invoices, L)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _batch_summary(ws, invoices, L):
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [3, 24, 18, 18, 18, 14]):
        ws.column_dimensions[col].width = w

    ws["B2"] = L["title"]
    ws["B2"].font = Font(name=FONT, size=16, bold=True, color=NAVY)

    # globalne sumy
    tot_net = tot_vat = tot_gross = 0.0
    n_items = 0
    currency = "PLN"
    for inv in invoices:
        f = inv.get("fields", {})
        tot_net += _num(_get(f, "net_amount")) or 0
        tot_vat += _num(_get(f, "tax_amount")) or 0
        tot_gross += _num(_get(f, "total_amount")) or 0
        n_items += len(inv.get("line_items", []) or [])
        currency = _get(f, "currency") or currency

    money_fmt = f'#,##0.00 "{currency}"'
    row = 4
    ws[f"B{row}"] = L["kpi"]
    ws[f"B{row}"].font = Font(name=FONT, size=10, bold=True, color=BLUE)
    row += 1
    kpis = [
        (L["count_inv"], len(invoices), "0"),
        (L["sum_net"], round(tot_net, 2), money_fmt),
        (L["sum_vat"], round(tot_vat, 2), money_fmt),
        (L["sum_gross"], round(tot_gross, 2), money_fmt),
        (L["count_items"], n_items, "0"),
    ]
    col = 2
    for label, val, fmt in kpis:
        c_lab = ws.cell(row=row, column=col, value=label)
        c_lab.font = Font(name=FONT, size=9, color="5A6782")
        c_lab.fill = PatternFill("solid", fgColor=LIGHT)
        c_lab.border = BORDER
        c_lab.alignment = Alignment(horizontal="center")
        c_val = ws.cell(row=row + 1, column=col, value=val)
        c_val.font = Font(name=FONT, size=13, bold=True, color=NAVY)
        c_val.border = BORDER
        c_val.alignment = Alignment(horizontal="center")
        if isinstance(val, (int, float)):
            c_val.number_format = fmt
        col += 1

    # kwoty wg sprzedawcy
    row += 3
    ws[f"B{row}"] = L["by_seller"]
    ws[f"B{row}"].font = Font(name=FONT, size=10, bold=True, color=BLUE)
    row += 1
    for i, h in enumerate([L["seller"], L["gross"]]):
        c = ws.cell(row=row, column=2 + i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    header_row = row
    row += 1

    by_seller: dict[str, float] = {}
    for inv in invoices:
        f = inv.get("fields", {})
        s = _seller_of(f)
        by_seller.setdefault(s, 0.0)
        by_seller[s] += _num(_get(f, "total_amount")) or 0

    data_start = row
    for seller, gross in sorted(by_seller.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=2, value=seller).border = BORDER
        ws.cell(row=row, column=2).font = Font(name=FONT, size=10)
        cg = ws.cell(row=row, column=3, value=round(gross, 2))
        cg.number_format = money_fmt
        cg.border = BORDER
        cg.font = Font(name=FONT, size=10)
        row += 1
    data_end = row - 1

    # wykres słupkowy: brutto wg sprzedawcy
    if by_seller:
        chart = BarChart()
        chart.title = L["chart_title"]
        chart.type = "bar"
        chart.legend = None
        cats = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
        vals = Reference(ws, min_col=3, min_row=header_row, max_row=data_end)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, f"B{row + 1}")


def _batch_invoices(ws, invoices, L):
    ws.sheet_view.showGridLines = False
    ws["A1"] = L["invoices_title"]
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)

    headers = [L["nr"], L["date"], L["seller"], L["buyer"],
               L["net"], L["vat"], L["gross"], L["currency"], L["file"]]
    widths = [16, 12, 26, 26, 14, 14, 14, 8, 22]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 4
    first = r
    tot_net = tot_vat = tot_gross = 0.0
    for inv in invoices:
        f = inv.get("fields", {})
        currency = _get(f, "currency") or "PLN"
        money_fmt = f'#,##0.00 "{currency}"'
        net = _num(_get(f, "net_amount")) or 0
        vat = _num(_get(f, "tax_amount")) or 0
        gross = _num(_get(f, "total_amount")) or 0
        tot_net += net; tot_vat += vat; tot_gross += gross
        vals = [
            _get(f, "invoice_number") or "—",
            _get(f, "issue_date") or "",
            _seller_of(f),
            _get(f, "buyer_name") or _get(f, "buyer") or "",
            net, vat, gross,
            currency,
            inv.get("source_file", ""),
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = BORDER
            cell.font = Font(name=FONT, size=10)
            if ci in (5, 6, 7):
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")
        if (r - first) % 2 == 1:
            for ci in range(1, 10):
                ws.cell(row=r, column=ci).fill = PatternFill("solid", fgColor=LIGHT)
        r += 1

    # wiersz RAZEM
    if invoices:
        ws.cell(row=r, column=4, value=L["total_row"]).font = Font(name=FONT, bold=True, color=NAVY)
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="right")
        for ci, v in [(5, tot_net), (6, tot_vat), (7, tot_gross)]:
            cell = ws.cell(row=r, column=ci, value=round(v, 2))
            cell.font = Font(name=FONT, bold=True, color=NAVY)
            cell.number_format = '#,##0.00'
            cell.fill = PatternFill("solid", fgColor=MID)
            cell.border = BORDER


def _batch_items(ws, invoices, L):
    ws.sheet_view.showGridLines = False
    ws["A1"] = L["items_title"]
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)

    headers = [L["nr"], L["desc"], L["qty"], L["unit_price"], L["vat_rate"], L["value_net"]]
    widths = [16, 34, 8, 14, 10, 15]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center" if i != 2 else "left")
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 4
    for inv in invoices:
        f = inv.get("fields", {})
        inv_no = _get(f, "invoice_number") or inv.get("source_file", "—")
        currency = _get(f, "currency") or "PLN"
        money_fmt = f'#,##0.00 "{currency}"'
        for it in inv.get("line_items", []) or []:
            if not isinstance(it, dict):
                continue
            qty = _num(it.get("quantity"))
            price = _num(it.get("unit_price"))
            net_val = round(qty * price, 2) if (qty is not None and price is not None) else _num(it.get("total"))
            vals = [inv_no, it.get("description", ""), qty if qty is not None else "",
                    price if price is not None else "", it.get("vat_rate", "") or "",
                    net_val if net_val is not None else ""]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.border = BORDER
                cell.font = Font(name=FONT, size=10)
                if ci in (4, 6):
                    cell.number_format = money_fmt
                if ci == 2:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="right" if ci != 1 else "left")
            r += 1


def build_batch_csv(invoices: list[dict]) -> bytes:
    """Zbiorczy CSV pod Power BI - jeden wiersz = jedna pozycja z każdej faktury."""
    from powerbi_export import build_powerbi_csv, COLUMNS
    import csv

    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=COLUMNS, extrasaction="ignore")
    writer.writeheader()

    for inv in invoices:
        # wykorzystujemy istniejący builder, ale zbieramy tylko wiersze danych
        single = build_powerbi_csv(inv).decode("utf-8")
        lines = single.split("\r\n")
        # pomijamy BOM+nagłówek (pierwsza linia) i puste
        for line in lines[1:]:
            if line.strip():
                out.write(line + "\r\n")

    return "\ufeff".encode("utf-8") + out.getvalue().encode("utf-8")
